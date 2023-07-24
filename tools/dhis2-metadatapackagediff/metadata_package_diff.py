import json
import chardet
import sys
import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_dataframe import get_as_dataframe, set_with_dataframe
from gspread_formatting import *
from gspread.exceptions import APIError
import time
import argparse
import re
import os

keys_not_owned = {
    'categories': ['items', 'categoryCombos'],
    'categoryCombos': ['categoryOptionCombos', 'items'],
    'categoryOptionCombos': ['sharing', 'shortName'],
    'categoryOptionGroupSets': ['shortName', 'items'],
    'categoryOptionGroups': ['groupSets'],
    'categoryOptions': ['categories', 'categoryOptionGroups', 'categoryOptionCombos'],
    'dataElements': ['dataSetElements', 'dataElementGroups'],
    'dataElementGroups': ['aggregationType'],
    'dataSets': ['formType', 'sections'],
    'indicatorTypes': ['sharing'],
    'indicators': ['indicatorGroups'],
    'maps': ['subscribed'],
    'sections': ['categoryCombos', 'sharing'],
    'userRoles': ['users'],
    'users': ['name', 'sharing'],
    'validationRules': ['validationRuleGroups', 'displayInstruction'],
    'visualizations': ['displayDomainAxisLabel', 'displayRangeAxisLabel', 'rangeAxisLabel', 'domainAxisLabel',
                       'displayTitle', 'subscribed']
}

def reindex(json_object, key):
    new_json = dict()
    for elem in json_object:
        key_value = elem[key]
        # elem.pop(key)
        new_json[key_value] = elem
    return new_json


def json_extract(obj, key):
    """Recursively fetch values from nested JSON."""
    arr = []

    def extract(obj, arr, key):
        """Recursively search for values of key in JSON tree."""
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, (dict, list)):
                    extract(v, arr, key)
                elif k == key:
                    arr.append(v)
        elif isinstance(obj, list):
            for item in obj:
                extract(item, arr, key)
        return arr

    values = extract(obj, arr, key)
    return values


# Convert json to dictionary
def json_to_dict(obj):
    separator = '.'

    result_dict = dict()
    key = ""

    skip_keys = ['translations', 'lastUpdated', 'lastUpdatedBy', 'createdBy', 'sharing.owner', 'href', 'access', 'created', 'allItems',
                 'displayName', 'displayDescription', 'displayNumeratorDescription', 'displayDenominatorDescription',
                 'displayFormName', 'displayShortName', 'dimension', 'dimensionType', 'isDefault', 'itemCount',
                 'user', 'dimensionItem', 'dimensionItemType', 'externalAccess', 'favorite', 'optionSetValue', 'periodOffset', 'publicAccess', 'externalAccess', 'userGroupAccesses'
    ]

    def scan(obj, result_dict, key):
        if key == "":
            s = ""
        else:
            s = separator
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k not in skip_keys:
                    if isinstance(v, (dict, list)):
                        scan(v, result_dict, key + s + k)
                    else:
                        final_key = key + s + k
                        if final_key not in result_dict:
                            result_dict[final_key] = v
                        else:
                            if isinstance(result_dict[final_key], list):
                                result_dict[final_key].append(v)
                            # convert into a list
                            else:
                                result_dict[final_key] = [result_dict[final_key], v]
                            # result_dict[final_key].sort()

        elif isinstance(obj, list):
            for item in obj:
                scan(item, result_dict, key)
        return result_dict

    values = scan(obj, result_dict, key)
    return values


def json_extract_nested_ids(obj, key):
    """Recursively fetch values from nested JSON."""
    arr = []

    def extract(obj, arr, key):
        """Recursively search for values of key in JSON tree."""
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k == key:
                    if isinstance(v, list):
                        for item in v:
                            arr.append(item["id"]) if item["id"] not in arr else arr
                    else:
                        arr.append(v)
                elif isinstance(v, (dict, list)):
                    extract(v, arr, key)
        elif isinstance(obj, list):
            for item in obj:
                extract(item, arr, key)
        return arr

    values = extract(obj, arr, key)
    return values


# Function to insert row in the dataframe
def insert_row(row_number, df, row_value):
    # Slice the upper half of the dataframe
    df1 = df[0:row_number]

    # Store the result of lower half of the dataframe
    df2 = df[row_number:]

    # Insert the row in the upper half dataframe
    # df1.loc[row_number] = row_value
    df1 = df1.append(row_value, ignore_index=True)

    # Concat the two dataframes
    df_result = pd.concat([df1, df2])

    # Reassign the index labels
    df_result.index = [*range(df_result.shape[0])]

    # Return the updated dataframe
    return df_result


def append_row_element(metaobj, df, type, operation, update = []):
    cols_to_add = ['id', 'name', 'lastUpdated', 'lastUpdatedBy']
    values = dict()
    if type not in df:
        df[type] = pd.DataFrame({}, columns=['operation', 'uid', 'name',
                                   'update_operation', 'update_key', 'update_diff', 'last_updated', 'updated_by'])

    for col in cols_to_add:
        if col in metaobj:
            values[col] = metaobj[col]
        else:
            values[col] = ""

    # CREATED -> Insert it at the beginning
    # DELETED -> Insert it at the end
    # UPDATED -> After the last CREATED or the first DELETED
    if operation == 'CREATED':
        insert_pos = 0
    elif operation == 'DELETED':
        insert_pos = df[type].shape[0]
    else:
        created_indexes = df[type].index[df[type]['operation'] == 'CREATED'].tolist()
        if len(created_indexes) > 0:
            insert_pos = created_indexes[-1]+1
        else:
            deleted_indexes = df[type].index[df[type]['operation'] == 'DELETED'].tolist()
            if len(deleted_indexes) > 0:
                insert_pos = deleted_indexes[0]
            else:
                # No CREATED no DELETED, just at the end
                insert_pos = df[type].shape[0]

    if operation != 'UPDATE' and len(update) == 0:
        df[type] = insert_row(insert_pos, df[type],
            {'uid': values['id'], 'name': values['name'], 'operation': operation,
             'update_operation': '', 'update_key': '', 'update_diff': '',
             'last_updated': values['lastUpdated'], 'updated_by': values['lastUpdatedBy']})
    else:
        if len(update) > 0:
            pos_increment = 0
            for upd in update:
                if pos_increment == 0:
                    df[type] = insert_row(insert_pos+pos_increment, df[type],
                        {'uid': values['id'], 'name': values['name'], 'operation': operation,
                         'update_operation': upd['update_operation'], 'update_key': upd['update_key'], 'update_diff': upd['update_diff'],
                         'last_updated': values['lastUpdated'], 'updated_by': values['lastUpdatedBy']})
                else:
                    df[type] = insert_row(insert_pos+pos_increment, df[type],
                        {'uid': '', 'name': '', 'operation': '',
                         'update_operation': upd['update_operation'], 'update_key': upd['update_key'], 'update_diff': upd['update_diff'],
                         'last_updated': '', 'updated_by': ''})
                pos_increment += 1
    return df


# Get element present in list2 but not in list1
def diff_list(list1, list2):
    result = []
    for elem in list2:
        if elem not in list1:
            result.append(elem)
    return result


def apply_conditional_format_to_ws(worksheet):

    condition_range = ['A2:C', 'A2:C', 'A2:C', 'D2:F', 'D2:F', 'D2:F']
    condition_color = [Color(0,1,0), Color(1,0,0), Color(0.11,0.56,1), Color(0,1,0), Color(1,0,0), Color(0.11,0.56,1)]
    custom_formula = ['=$A2="CREATED"','=$A2="DELETED"','=$A2="UPDATED"','=$D2="CREATED"','=$D2="DELETED"','=$D2="UPDATED"']

    # Get number of rows
    number_of_rows = worksheet.row_count

    # Get the rules list to append new ones
    rules = get_conditional_format_rules(worksheet)
    rules.clear()

    for i in range(0,len(condition_range)):
        rule = ConditionalFormatRule(
            ranges=[GridRange.from_a1_range(condition_range[i]+str(number_of_rows+1), worksheet)],
            booleanRule=BooleanRule(
                condition=BooleanCondition('CUSTOM_FORMULA', [custom_formula[i]]),
                format=CellFormat(backgroundColor=condition_color[i])
            )
        )

        rules.append(rule)

    rules.save()


if __name__ == '__main__':


    my_parser = argparse.ArgumentParser(prog='metadata_package_diff',
                                        description='Compare to packages given by 2 json files and uploads the result to Google Spreadsheets',
                                        epilog="python metadata_package_diff.py old_file.json new_file.json ""My comparison"" --share_with=johndoe@dhis2.org",
                                        formatter_class=argparse.RawDescriptionHelpFormatter)
    my_parser.add_argument('old_json', metavar='old_json', type=str,
                           help='Previous json file')
    my_parser.add_argument('new_json', metavar='new_json', type=str,
                           help='New json file')
    my_parser.add_argument('sheet_name', metavar='sheet_name', type=str,
                           help='Title of the spreadsheet to use or create (if it exists in the workspace, it gets updated)')
    my_parser.add_argument('-ug', '--use_gspread', action="store_true", dest='use_gspread',
                           help='Upload result to Google Spreadsheet (requires token)')
    my_parser.set_defaults(use_gspread=False)
    my_parser.add_argument('-sw', '--share_with', action="append", metavar='email', nargs=1,
                           help='email address to share the generated spreadsheet with as OWNER. '
                                'Eg: --share_with=peter@dhis2.org')
    args = my_parser.parse_args()
    package_file1 = args.old_json
    package_file2 = args.new_json

    upload_to_gspread = args.use_gspread

    sh_name = args.sheet_name

    if not os.path.exists(package_file1):
        print("The file " + package_file1 + " could not be found")
        exit(1)
    if not os.path.exists(package_file2):
        print("The file " + package_file2 + " could not be found")
        exit(1)
    if upload_to_gspread:
        if args.share_with is not None and len(args.share_with) > 0:
            for param in args.share_with:
                if not (re.search('^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$', param[0])):
                    print("The email address " + param[0] + " is not valid")
                    exit(1)

        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        google_spreadshseet_credentials = 'dummy-data-297922-97b90db83bdc.json'
        try:
            f = open(google_spreadshseet_credentials)
        except IOError:
            print("Please provide file with google spreadsheet credentials")
            exit(1)
        else:
            credentials = ServiceAccountCredentials.from_json_keyfile_name(google_spreadshseet_credentials, scope)

        gc = gspread.authorize(credentials)
        mode='update'
        sheet1_still_there = False
        try:
            gs = gc.open(sh_name)
        except gspread.SpreadsheetNotFound:
            mode='create'
            sheet1_still_there = True
            gs = gc.create(sh_name)
            pass

    df = dict()
    delete_payload = dict()

    enc = chardet.detect(open(package_file1, 'rb').read())['encoding']
    with open(package_file1, 'r', encoding=enc) as json_file:
        metadata1 = json.load(json_file)
    with open(package_file2, 'r', encoding=enc) as json_file:
        metadata2 = json.load(json_file)

    ###### Check if new metadata types have been deleted/created
    keys1 = list(metadata1.keys())
    keys2 = list(metadata2.keys())

    # Consolidate all keys into one list and order it alphabetically
    keys = list(dict.fromkeys(keys1 + keys2))
    keys.sort()
    for key in keys:
        # All elements have been removed
        if key in metadata1 and key not in metadata2:
            for element in metadata1[key]:
                df = append_row_element(element, df, key, 'DELETED')
        # All elements have been created
        elif key in metadata2 and key not in metadata1:
            for element in metadata2[key]:
                df = append_row_element(element, df, key, 'CREATED')
        # Look at individual elements
        else:
            ##### Check that a single metadata type and look for new/deleted/updated elements
            if isinstance(metadata1[key], list) and isinstance(metadata2[key], list):
                # Convert a list into a dictionary indexed by uid
                metaobj1 = reindex(metadata1[key], 'id')
                metaobj2 = reindex(metadata2[key], 'id')
                uids1 = list(metaobj1.keys())
                uids2 = list(metaobj2.keys())

                # Consolidate all keys into one list and order it alphabetically
                uids = list(dict.fromkeys(uids1 + uids2))
                uids.sort()

                for uid in uids:
                    if uid in metaobj1 and uid not in metaobj2:
                        df = append_row_element(metaobj1[uid], df, key, 'DELETED')
                        if key not in delete_payload:
                            delete_payload[key] = list()
                        delete_payload[key].append(metaobj1[uid])
                    # All elements have been created
                    elif uid in metaobj2 and uid not in metaobj1:
                        df = append_row_element(metaobj2[uid], df, key, 'CREATED')
                    # This element may have been updated
                    else:
                        #### Check for created/deleted/updated fields within same elements
                        # Try to use here the diff tool metaobj1[uid] VS metaobj2[uid]
                        dict1 = json_to_dict(metaobj1[uid])
                        dict2 = json_to_dict(metaobj2[uid])

                        k1 = list(dict1.keys())
                        k2 = list(dict2.keys())

                        # Consolidate all keys into one list and order it alphabetically
                        elem_keys = list(dict.fromkeys(k1 + k2))
                        elem_keys.sort()

                        update_list = []
                        for k in elem_keys:
                            if k in dict1 and k not in dict2:
                                if key not in keys_not_owned or not [item for item in keys_not_owned[key] if item in k]:
                                    if isinstance(dict1[k], list):
                                        #update_list.append("DELETED|" + k + " : " + 'list(' + str(len(dict1[k])) + ')')
                                        update_list.append({"update_operation":"DELETED",
                                                            "update_key": k + " : " + 'list(' + str(len(dict1[k])) + ')',
                                                            "update_diff": ""})
                                    else:
                                        #update_list.append("DELETED|" + k + " : " + str(dict1[k]))
                                        update_list.append({"update_operation":"DELETED",
                                                            "update_key": k + " : " + str(dict1[k]),
                                                            "update_diff": ""})
                            # All elements have been created
                            elif k in dict2 and k not in dict1:
                                if isinstance(dict2[k], list):
                                    #update_list.append("CREATED|" + k + " : " + 'list(' + str(len(dict2[k])) + ')')
                                    update_list.append({"update_operation":"CREATED",
                                                        "update_key": k + " : " + 'list(' + str(len(dict2[k])) + ')',
                                                        "update_diff": ""})
                                else:
                                    #update_list.append("CREATED|" + k + " : " + str(dict2[k]))
                                    update_list.append({"update_operation":"CREATED",
                                                        "update_key": k + " : " + str(dict2[k]),
                                                        "update_diff": ""})
                            # This element may have been updated
                            else:
                                if isinstance(dict1[k], list) and isinstance(dict2[k], list):
                                    if dict1[k].sort() != dict2[k].sort():
                                        msg = ""
                                        # Get deleted elements in second list
                                        deleted = diff_list(dict2[k], dict1[k])
                                        if len(deleted) > 0:
                                            msg += "DEL: " + str(deleted)
                                        # Get added elements in second list
                                        added = diff_list(dict1[k], dict2[k])
                                        if len(added) > 0:
                                            if msg != "":
                                                msg += " "
                                            msg += "ADD: " + str(added)
                                        #update_list.append("UPDATED|" + k + "|" + msg)
                                        update_list.append({"update_operation": "UPDATED",
                                                            "update_key": k,
                                                            "update_diff": msg})

                                else:
                                    if dict1[k] != dict2[k]:
                                        #update_list.append("UPDATED|" + k + "|" + str(dict1[k]) + " -> " + str(dict2[k]))
                                        # There is a restriction of 50000 maximum characters in a cell so we cannot do the update of
                                        # things like htmlCode
                                        if len(str(dict1[k]) + " -> " + str(dict2[k])) > 50000:
                                            update_list.append({"update_operation": "UPDATED",
                                                                "update_key": k,
                                                                "update_diff": "LENGHT_OF_COMPARISON_IS_TOO_BIG_TO_BE_DISPLAYED"})
                                        else:
                                            update_list.append({"update_operation": "UPDATED",
                                                                "update_key": k,
                                                                "update_diff": str(dict1[k]) + " -> " + str(dict2[k])})
                        if len(update_list) > 0:
                            df = append_row_element(metaobj2[uid], df, key, 'UPDATED', update_list)

            # It is most likely the PACKAGE label
            else:
                if metadata1[key] != metadata2[key]:
                    if key not in df:
                        df[key] = pd.DataFrame({})
                    df[key] = insert_row(0, df[key],
                        {'uid': '', 'name': '', 'operation': 'UPDATED',
                         'update_operation': 'UPDATED', 'update_key': '',
                         'update_diff': str(metadata1[key]) + " -> " + str(metadata2[key]),
                         'last_updated': '', 'updated_by': ''})
                    # df[key] = pd.DataFrame(
                    #     {'uid': '', 'name': '', 'operation': 'UPDATED',
                    #      'update_operation': 'UPDATED', 'update_key': '',
                    #      'update_diff': str(metadata1[key]) + " -> " + str(metadata2[key]),
                    #      'last_updated': '', 'updated_by': ''})


    # dateTimeObj = datetime.now()
    # timestampStr = dateTimeObj.strftime("%d-%b-%Y_%H-%M-%S")
    # export_csv = df.to_csv(r'./comparison_' + timestampStr + '.csv', index=None, header=True)

    # Loop though every metadata type

    if delete_payload != {}:
        with open('package_diff_delete.json', 'w', encoding='utf8') as file:
            file.write(json.dumps(delete_payload, indent=4, sort_keys=True, ensure_ascii=False))
        file.close()

    writer = pd.ExcelWriter(sh_name + '.xlsx', engine='openpyxl')
    for metadata_type in df:
        print('Processing ' + metadata_type)

        df[metadata_type].to_excel(writer, sheet_name=metadata_type, index=False, header=True)

        # Get the worksheet
        ws = writer.sheets[metadata_type]

        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Freeze the first row
        ws.freeze_panes = ws['A2']

        index = 1
        for column in df[metadata_type].columns:
            width = df[metadata_type][column].str.len().max()
            if width < 10:
                width = 10
            ws.column_dimensions[get_column_letter(index)].width = width
            index += 1


        fill_red = openpyxl.styles.PatternFill(start_color='00FF0000', end_color='00FF0000',
                                           fill_type='solid')
        fill_blue = openpyxl.styles.PatternFill(start_color='001C8FFF', end_color='001C8FFF',
                                           fill_type='solid')
        fill_green = openpyxl.styles.PatternFill(start_color='0000FF00', end_color='0000FF00',
                                           fill_type='solid')
        rule1 = FormulaRule(formula=['=$A2="CREATED"'], stopIfTrue=True, fill=fill_green)
        rule2 = FormulaRule(formula=['=$A2="UPDATED"'], stopIfTrue=True, fill=fill_blue)
        rule3 = FormulaRule(formula=['=$A2="DELETED"'], stopIfTrue=True, fill=fill_red)
        rule4 = FormulaRule(formula=['=$D2="CREATED"'], stopIfTrue=True, fill=fill_green)
        rule5 = FormulaRule(formula=['=$D2="UPDATED"'], stopIfTrue=True, fill=fill_blue)
        rule6 = FormulaRule(formula=['=$D2="DELETED"'], stopIfTrue=True, fill=fill_red)

        # Add the conditional formatting rules to the worksheet
        ws.conditional_formatting.add(get_column_letter(1) + '2:' + get_column_letter(3) + str(ws.max_row), rule1)
        ws.conditional_formatting.add(get_column_letter(1) + '2:' + get_column_letter(3) + str(ws.max_row), rule2)
        ws.conditional_formatting.add(get_column_letter(1) + '2:' + get_column_letter(3) + str(ws.max_row), rule3)

        ws.conditional_formatting.add(get_column_letter(4) + '2:' + get_column_letter(6) + str(ws.max_row), rule4)
        ws.conditional_formatting.add(get_column_letter(4) + '2:' + get_column_letter(6) + str(ws.max_row), rule5)
        ws.conditional_formatting.add(get_column_letter(4) + '2:' + get_column_letter(6) + str(ws.max_row), rule6)

    writer.save()


    if upload_to_gspread:
        for metadata_type in df:
            # if metadata_type not in ['optionSets', 'dataElements', 'visualizations']:
            #     continue
            successful = False
            while not successful:
                print('Processing ' + metadata_type)
                metadata_type_ws_exists = True
                try:
                    gs.worksheet(metadata_type)
                except gspread.WorksheetNotFound:
                    metadata_type_ws_exists = False

                try:
                    if mode == 'create' or not metadata_type_ws_exists:
                        if sheet1_still_there:
                            ws = gs.sheet1
                            ws.update_title(metadata_type)
                            sheet1_still_there = False
                        else:
                            ws = gs.add_worksheet(title=metadata_type, rows=df[metadata_type].shape[0], cols=df[metadata_type].shape[1])
                    else:
                        ws = gs.worksheet(metadata_type)

                    ws.clear()
                    set_with_dataframe(worksheet=ws, dataframe=df[metadata_type], include_index=False,
                                       include_column_header=True, resize=True)

                    ws.format('A1:H1', {'textFormat': {'bold': True}})
                    set_frozen(ws, rows=1)
                    apply_conditional_format_to_ws(ws)
                except APIError as e:
                    # Temporary fix for write requests per minute per user per project 60
                    # This could be improved by using batch requests
                    result = e.args[0]
                    if result['code'] == 429:
                        print('Quota exceeded, waiting 1min before retrying')
                        time.sleep(60)
                        pass
                    else:
                        print('UNHANDLED ERROR IN THE API REQUEST: ' + str(result))
                        exit()
                else:
                    successful = True

        gs.share('manuel@dhis2.org', perm_type='user', role='writer')
        if args.share_with is not None:
            for email in args.share_with:
                gs.share(email[0], perm_type='user', role='writer')

        google_spreadsheet_url = "https://docs.google.com/spreadsheets/d/%s" % gs.id
        print('Google spreadsheet created/updated here: ' + google_spreadsheet_url)


